import tkinter as tk
from tkinter import ttk, filedialog as fd, messagebox as mb
import traceback
import os
import threading
import time

from src.read_excel import read_to_xml
from src.generate_excel_template import generate_standard_templates
from src.dfs_schema import get_project_root, get_XML_schema

"""
Command:
pyinstaller -F -w -n xls2xml --distpath app gui.py
"""


class CollapsibleFrame(ttk.Frame):
    def __init__(self, parent, text, **kwargs):
        super().__init__(parent, **kwargs)

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        self.header = ttk.Button(self, text=f"▶ {text}", command=self.toggle)
        self.header.grid(row=0, column=0, sticky="ew")

        self.content_frame = ttk.Frame(self, padding=5)
        self.content_frame.grid(row=1, column=0, sticky="nsew")

        self.is_collapsed = True
        self.toggle()

    def toggle(self):
        if self.is_collapsed:
            self.header.config(text=f"▼ {self.header['text'][2:]}")
            self.content_frame.grid()
        else:
            self.header.config(text=f"▶ {self.header['text'][2:]}")
            self.content_frame.grid_remove()

        self.is_collapsed = not self.is_collapsed


class Xls2XmlApp(tk.Tk):
    """A simple GUI for converting Excel files to XML."""

    def __init__(self):
        super().__init__()

        # --- Constants and Configuration ---
        self.TITLE = 'xls2xml'
        self.DEFAULT_DIMENSIONS = '800x500'
        self.FILETYPES = (('All Excel files', '*.xl*'), ('All files', '*.*'))
        self.OUTPUT_FILETYPES = (('XML files', '*.xml'),)
        self.POSSIBLE_SHEETS = (
            'bodemkundigeopbouw', 'bodemlocatie', 'bodemlocatieclassificatie', 'bodemmonster', 'bodemobservatie',
            'bodemsite', 'boring', 'filter', 'filterdebietmeter', 'filtermeting', 'grondmonster', 'grondwaterlocatie',
            'interpretaties', 'monster', 'observatie', 'opdracht', 'sondering')

        self.SCHEMAS = {}
        self.sheet_checkboxes = {}

        # --- Main Window Setup ---
        self.title(self.TITLE)
        self.geometry(self.DEFAULT_DIMENSIONS)
        self.resizable(True, True)

        self._setup_variables()
        self._setup_widgets()
        self._setup_menu()
        self._setup_layout()
        self._toggle_sheets()  # Initialize sheet checkboxes as disabled

    def _setup_variables(self):
        """Initializes application variables."""
        self.input_filename = tk.StringVar()
        self.output_filename = tk.StringVar()
        self.omgeving = tk.StringVar(value='productie')
        self.all_sheets_var = tk.BooleanVar(value=True)  # Variable for the 'Automatic' checkbox
        self.sheet_vars = {}

    def _setup_widgets(self):
        """Creates and configures all GUI widgets."""
        self.input_entry = ttk.Entry(self, width=100, textvariable=self.input_filename)
        self.input_button = ttk.Button(self, text='Select Input File', command=self.select_file)

        self.output_entry = ttk.Entry(self, width=100, textvariable=self.output_filename)
        self.output_button = ttk.Button(self, text='Select Output File', command=self.save_as_file)

        self.omgeving_label = ttk.Label(self, text='Environment:')
        self.omgeving_menu = ttk.OptionMenu(self, self.omgeving, 'productie', 'oefen', 'ontwikkel')

        self.sheets_label = ttk.Label(self, text='Sheets:')
        self.automatic_checkbox = ttk.Checkbutton(self, text='Automatic (All Sheets)', variable=self.all_sheets_var,
                                                  command=self._toggle_sheets)

        self.run_button = ttk.Button(self, text='Run Conversion', command=self.run_xls2xml)
        self.status_label = ttk.Label(self, text='Ready.', relief=tk.SUNKEN, anchor=tk.W)
        self.progressbar = ttk.Progressbar(self, orient=tk.HORIZONTAL, length=400, mode='determinate')

    def _setup_menu(self):
        """Creates the application menu bar."""
        menubar = tk.Menu(self)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Generate Templates", command=self.generate_templates)
        menubar.add_cascade(label="File", menu=filemenu)
        self.config(menu=menubar)

    def _setup_layout(self, n=0):
        """Lays out the widgets using the grid manager."""
        self.input_entry.grid(row=0, column=0, pady=10, padx=(10, 5), sticky='NESW', columnspan=2)
        self.input_button.grid(row=0, column=2, pady=10, padx=(5, 10), sticky='NESW')
        self.output_entry.grid(row=1, column=0, pady=10, padx=(10, 5), sticky='NESW', columnspan=2)
        self.output_button.grid(row=1, column=2, pady=10, padx=(5, 10), sticky='NESW')

        self.omgeving_label.grid(row=2, column=0, pady=10, padx=10, sticky='W')
        self.omgeving_menu.grid(row=2, column=1, pady=10, padx=10, sticky='NESW')

        self.sheets_label.grid(row=3, column=0, pady=5, padx=10, sticky='W')
        self.automatic_checkbox.grid(row=3, column=1, pady=5, padx=10, sticky='W', columnspan=2)

        self.run_button.grid(row=5 + n, columnspan=3, pady=20, padx=10, sticky='ESW')
        self.progressbar.grid(row=6 + n, columnspan=3, pady=5, padx=10, sticky='NESW')
        self.status_label.grid(row=7 + n, columnspan=3, pady=(5, 10), padx=10, sticky='NESW')

        # Configure row and column weights for resizing
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=0)

        for i in range(n):
            self.grid_rowconfigure(4 + i, weight=0)

        self.grid_rowconfigure(4 + n, weight=1)

    def _get_sheet_names(self, file_path):
        """
        Retrieves sheet names from an Excel file.
        This function needs to be implemented using a library like openpyxl.
        """
        try:
            # Example using openpyxl
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            mb.showerror("Error", f"Failed to read sheet names from the file.\nError: {e}")
            return []

    def _create_sheet_checkboxes(self, sheets):
        """Dynamically creates checkboxes for each sheet."""
        # Clear existing checkboxes
        for widget in self.sheet_checkboxes.values():
            widget.destroy()
        self.sheet_checkboxes = {}
        self.sheet_vars = {}

        row_start = 4
        col = 0
        max_cols = 3  # Adjust as needed

        for i, sheet_name in enumerate(sheets):
            var = tk.BooleanVar(value=False)
            self.sheet_vars[sheet_name] = var
            chk_btn = ttk.Checkbutton(self, text=sheet_name, variable=var, command=self._update_automatic)
            self.sheet_checkboxes[sheet_name] = chk_btn

            row = row_start + (i // max_cols)
            col = i % max_cols
            chk_btn.grid(row=row, column=col, padx=10, pady=5, sticky='NW')

        self._setup_layout(len(sheets) // max_cols)

        self._toggle_sheets()  # Update initial state based on 'Automatic'

    def _toggle_sheets(self):
        """Enables/disables individual sheet checkboxes based on 'Automatic' state."""
        is_automatic = self.all_sheets_var.get()
        for sheet_name, chk_btn in self.sheet_checkboxes.items():
            if is_automatic:
                chk_btn.config(state='disabled')
                self.sheet_vars[sheet_name].set(False)
            else:
                chk_btn.config(state='enabled')

    def _update_automatic(self):
        """Updates the 'Automatic' checkbox state based on individual sheet selections."""
        if all(self.sheet_vars[name].get() for name in self.sheet_vars):
            self.all_sheets_var.set(True)
        else:
            self.all_sheets_var.set(False)
        self._toggle_sheets()

    def select_file(self):
        """Opens a file dialog to select the input Excel file."""
        filename = fd.askopenfilename(
            title='Select Input Excel File',
            initialdir='./',
            filetypes=self.FILETYPES
        )
        if filename:
            self.input_filename.set(filename)

            # Suggest an output filename based on the input

            base, _ = os.path.splitext(filename)
            default_output = f'{base}.xml'
            self.output_filename.set(default_output)

            # Get and display sheet names
            sheets = [s for s in self._get_sheet_names(filename) if s in self.POSSIBLE_SHEETS]
            self._create_sheet_checkboxes(sheets)

    def save_as_file(self):
        """Opens a file dialog to save the output XML file."""
        filename = fd.asksaveasfilename(
            defaultextension='.xml',
            initialdir='./dist',
            title='Save Output XML File',
            filetypes=self.OUTPUT_FILETYPES
        )
        if filename:
            self.output_filename.set(filename)

    def run_xls2xml(self):
        """Performs the conversion in a separate thread."""
        input_path = self.input_filename.get()
        output_path = self.output_filename.get()

        if not input_path or not output_path:
            mb.showwarning("Missing Information", "Please select both an input and an output file.")
            return

        selected_sheets = []
        if not self.all_sheets_var.get():
            selected_sheets = [name for name, var in self.sheet_vars.items() if var.get()]
            if not selected_sheets:
                mb.showwarning("No Sheets Selected", "Please select at least one sheet or the 'Automatic' option.")
                return

        self.run_button.config(state=tk.DISABLED)
        self.status_label.config(text='Preparing to convert...')

        # Start the conversion in a new thread
        conversion_thread = threading.Thread(
            target=self._perform_conversion,
            args=(input_path, output_path, selected_sheets),
            daemon=True
        )
        conversion_thread.start()

    def _perform_conversion(self, input_path, output_path, sheets_to_convert):
        """Performs the actual conversion (runs in a separate thread)."""
        self.status_label.config(text='Converting your xlsx to XML...')
        self.progressbar.start()  # Use indeterminate mode for a generic loading animation

        try:
            # Check if schema is already loaded
            if self.omgeving.get() not in self.SCHEMAS:
                self.status_label.config(text=f'Loading schema for {self.omgeving.get()}...')
                self.update_idletasks()
                self.SCHEMAS[self.omgeving.get()] = get_XML_schema(self.omgeving.get())

            self.status_label.config(text=f'Converting your xls to xml...')
            self.update_idletasks()

            rapport = read_to_xml(input_path, output_path, xsd_source=self.omgeving.get(),
                                  project_root=get_project_root(), xml_schema=self.SCHEMAS[self.omgeving.get()],
                                  sheets=sheets_to_convert)

            for _ in range(200):
                time.sleep(0.01)
                self.update_idletasks()
            self.progressbar.stop()
            self.progressbar['value'] = 0

            self.show_custom_message('Conversion Complete!',
                                     f'The conversion was completed!\nYour XML file can be found at: {output_path}\n' +
                                     f'Logs:\n{rapport.get_error_rapport()}'
                                     )

            self.status_label.config(text='Conversion complete.')
        except Exception as e:
            self.show_custom_message('Conversion Failed!',
                                     'The conversion to XML has failed.\n'
                                     f'Reason: {e}\n'
                                     'Please check the console for more details.')
            print(f"Detailed Error:\n{traceback.format_exc()}")
            self.status_label.config(text='Error: Conversion failed.')
        finally:
            self.progressbar.stop()
            self.progressbar['value'] = 0
            self.run_button.config(state=tk.NORMAL)

    def generate_templates(self):
        """Generates standard Excel templates."""
        try:
            generate_standard_templates(project_root=get_project_root(), mode='online')
            mb.showinfo('Templates Generated', 'Templates were successfully generated.')
        except Exception as e:
            self.show_custom_message('Update Failed',
                                     "Your version is outdated and can't generate the most recent templates.\n"
                                     f"Reason: {e}")

    def show_custom_message(self, title, message):
        """
        Creates a custom dialog box with a title and a message that can be copied.
        This function is reusable for both success messages and error reports.
        """

        dialog = tk.Toplevel()
        dialog.title(title)
        dialog.geometry("600x400")  # A bit bigger for error messages
        dialog.resizable(True, True)

        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill="both", expand=True)

        text_widget = tk.Text(frame, wrap="word", padx=5, pady=5)
        text_widget.insert("1.0", message)
        text_widget.config(state="disabled")

        scrollbar = ttk.Scrollbar(frame, command=text_widget.yview)
        text_widget.config(yscrollcommand=scrollbar.set)

        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        copy_button = ttk.Button(dialog, text="Copy to Clipboard",
                                 command=lambda: dialog.clipboard_clear() or dialog.clipboard_append(message))
        copy_button.pack(pady=10)

        dialog.grab_set()
        dialog.focus_set()
        dialog.wait_window()

    def _show_logging_window(self, validator_object):
        log_window = tk.Toplevel()
        log_window.title("Conversion Log")
        log_window.geometry("800x600")

        main_frame = ttk.Frame(log_window, padding=10)
        main_frame.pack(fill="both", expand=True)

        # Main Scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Global log message
        global_message_frame = CollapsibleFrame(scrollable_frame, text="Global Log")
        global_message_frame.pack(fill="x", pady=5)

        # You'll need to get the global message from your Validator object
        global_message = "Overall conversion status: Success"
        ttk.Label(global_message_frame.content_frame, text=global_message, wraplength=750).pack(fill="x")

        # Loop through each sheet's log data
        # Assuming validator_object has a structure like {'Sheet1': [logs], 'Sheet2': [logs]}
        for sheet_name, logs in validator_object.items():
            sheet_frame = CollapsibleFrame(scrollable_frame, text=f"Log for '{sheet_name}'")
            sheet_frame.pack(fill="x", pady=5)

            log_text = tk.Text(sheet_frame.content_frame, height=10, state='disabled')
            log_text.pack(fill="both", expand=True)

            log_text.config(state='normal')
            for log_entry in logs:
                log_text.insert(tk.END, f"{log_entry}\n")
            log_text.config(state='disabled')

        log_window.grab_set()
        log_window.focus_set()
        log_window.wait_window()


if __name__ == '__main__':
    app = Xls2XmlApp()
    app.mainloop()
